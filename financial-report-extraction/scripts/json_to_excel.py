import os
from typing import Any, Dict, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalize_column_name(value: Any) -> str:
    text = _normalize_text(value)
    aliases = {
        "期初": "年初数",
        "期初值": "年初数",
        "年初": "年初数",
        "期末": "期末数",
        "期末值": "期末数",
        "本月": "本月数",
        "本月值": "本月数",
        "本年累计": "本年累计数",
        "本年累计值": "本年累计数",
        "本年": "本年金额",
        "本年金额": "本年金额",
    }
    return aliases.get(text, text)


def _as_list(value: Any) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple) or isinstance(value, set):
        return list(value)
    return [value]


def _collect_highlight_targets(
    abnormal_items: Optional[Any],
) -> Tuple[Set[str], Dict[str, Set[str]], Dict[str, str]]:
    row_names: Set[str] = set()
    cell_targets: Dict[str, Set[str]] = {}
    reasons: Dict[str, str] = {}

    for item in _as_list(abnormal_items):
        if isinstance(item, dict):
            names = []
            for key in ("科目名称", "涉及科目", "异常数据", "项目", "科目", "结果科目"):
                for value in _as_list(item.get(key)):
                    text = _normalize_text(value)
                    if text:
                        names.append(text)

            columns = set()
            for key in ("字段", "列名", "金额字段", "异常字段", "期间"):
                for value in _as_list(item.get(key)):
                    column_name = _normalize_column_name(value)
                    if column_name:
                        columns.add(column_name)

            reason = _normalize_text(item.get("失败原因") or item.get("原因") or item.get("说明"))
            for name in names:
                row_names.add(name)
                if columns:
                    cell_targets.setdefault(name, set()).update(columns)
                if reason:
                    reasons[name] = reason
        else:
            text = _normalize_text(item)
            if text:
                row_names.add(text)

    return row_names, cell_targets, reasons


def _matches_subject(row_name: str, target_name: str) -> bool:
    if not row_name or not target_name:
        return False
    return row_name == target_name or row_name in target_name or target_name in row_name


def json_data_to_excel(
    vlm_text: str,
    original_filename: str,
    json_out_dir: Optional[str] = None,
    sheet_name: str = "Sheet1",
    abnormal_items: Optional[Any] = None,
):
    """
    将 JSON 数据转换为 Excel 文件。

    :param vlm_text: JSON 格式的字符串，包含表名、编制单位、日期、单位、表格数据
    :param original_filename: 原始文件名，用于构建输出路径
    :param json_out_dir: Excel 文件输出目录，默认当前工作目录下的 result 目录
    :param sheet_name: 工作表名称，默认为 "Sheet1"
    :param abnormal_items: 勾稽失败异常项列表，用于高亮涉及的科目行或金额单元格
    """
    import json

    # 解析 JSON 数据
    data = json.loads(vlm_text)
    original_stem = os.path.splitext(os.path.basename(original_filename))[0]

    if json_out_dir is None:
        json_out_dir = os.path.join(os.getcwd(), "result", original_stem)

    # 确保输出目录存在
    os.makedirs(json_out_dir, exist_ok=True)

    # 构建输出文件路径
    output_path = os.path.join(json_out_dir, f"{original_stem}_origin.xlsx")

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws.title = sheet_name

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    abnormal_row_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    abnormal_cell_fill = PatternFill(
        start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
    )
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    abnormal_font = Font(color="9C0006")

    # 写入表头信息（表名、编制单位、日期、单位）
    table_name = data.get("表名", "")
    company_name = data.get("编制单位", "")
    date_str = data.get("日期", "")
    unit = data.get("单位", "")

    # 第一行：表名
    ws.cell(row=1, column=1, value=table_name)
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    # 第二行：编制单位、日期、单位
    ws.cell(row=2, column=1, value=f"编制单位：{company_name}")
    ws.cell(row=2, column=3, value=f"日期：{date_str}")
    ws.cell(row=2, column=5, value=f"单位：{unit}")

    # 获取表格数据
    table_data = data.get("表格数据", [])
    if not table_data:
        raise ValueError("JSON 数据中未找到表格数据")

    # 确定列名（从第一条数据中获取）
    first_row = table_data[0]
    columns = list(first_row.keys())
    abnormal_rows, abnormal_cells, abnormal_reasons = _collect_highlight_targets(
        abnormal_items
    )

    # 计算需要合并的列数（用于表名行）
    num_cols = len(columns)
    if num_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    # 第三行：列标题
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # 写入数据行
    for row_idx, row_data in enumerate(table_data, start=4):
        account_name = _normalize_text(row_data.get("科目名称", ""))
        matched_targets = [
            target for target in abnormal_rows if _matches_subject(account_name, target)
        ]
        matched_cell_columns = set()
        abnormal_reason = ""
        for target in matched_targets:
            matched_cell_columns.update(abnormal_cells.get(target, set()))
            abnormal_reason = abnormal_reason or abnormal_reasons.get(target, "")

        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if matched_targets:
                highlight_cell = not matched_cell_columns or col_name in matched_cell_columns
                cell.fill = abnormal_cell_fill if highlight_cell else abnormal_row_fill
                if highlight_cell:
                    cell.font = abnormal_font
                    if abnormal_reason:
                        cell.comment = Comment(abnormal_reason, "financial-report-extraction")

    # 调整列宽
    for col_idx in range(1, num_cols + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and not isinstance(cell, MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)
    return output_path
